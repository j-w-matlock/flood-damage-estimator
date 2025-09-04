import streamlit as st
import os
import tempfile
import pandas as pd
import rasterio
from utils.processing import (
    process_flood_damage,
    run_monte_carlo,
    constant_depth_array,
    sanitize_label,
)
from utils.crop_definitions import (
    CROP_DEFINITIONS,
    CROP_GROWING_SEASONS,
    DEFAULT_GROWING_SEASON,
)
import matplotlib.pyplot as plt
from collections import Counter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(layout="wide")
st.title("🌾 Agricultural Flood Damage Estimator")

# Session state init
for key in [
    "result_path",
    "summaries",
    "diagnostics",
    "crop_path",
    "depth_inputs",
    "label_map",
    "label_metadata",
    "crop_inputs",
    "temp_dir",
    "upload_dir",
]:
    if key not in st.session_state:
        if key == "crop_inputs":
            st.session_state[key] = {}
        else:
            st.session_state[key] = None

upload_dir = st.session_state.get("upload_dir")
if upload_dir:
    upload_dir.cleanup()
st.session_state["upload_dir"] = tempfile.TemporaryDirectory()


def cleanup_temp_dir():
    tmp = st.session_state.get("temp_dir")
    if tmp:
        tmp.cleanup()
    st.session_state["temp_dir"] = None


def save_upload(uploaded, suffix):
    upload_dir = st.session_state.get("upload_dir")
    if upload_dir is None:
        upload_dir = tempfile.TemporaryDirectory()
        st.session_state["upload_dir"] = upload_dir
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=suffix, dir=upload_dir.name
    ) as tmp:
        tmp.write(uploaded.read())
        return tmp.name


# Reset
if st.sidebar.button("🔁 Reset App"):
    cleanup_temp_dir()
    upload_dir = st.session_state.get("upload_dir")
    if upload_dir:
        upload_dir.cleanup()
    for key in list(st.session_state.keys()):
        del st.session_state[key]
    st.rerun()

# Sidebar Inputs
st.sidebar.header("🛠️ Settings")
mode = st.sidebar.radio(
    "Select Analysis Mode:",
    ["Direct Damages", "Monte Carlo Simulation"],
    help="Choose whether to run a straightforward flood loss calculation (Direct Damages) or include uncertainty using random simulations (Monte Carlo Simulation)",
)
crop_file = st.sidebar.file_uploader(
    "🌾 USDA CropScape Raster",
    type=["tif", "img"],
    help="Upload a CropScape raster that defines crop type per pixel",
)
depth_files = st.sidebar.file_uploader(
    "🌊 Flood Depth Grids",
    type=["tif"],
    accept_multiple_files=True,
    help="Upload one or more flood depth raster files (in feet)",
)
use_uniform_depth = st.sidebar.checkbox(
    "🏞️ Uniform Flood Depth",
    value=False,
    help="Duplicate the crop raster and fill with a user specified constant depth",
)
uniform_depth_ft = st.sidebar.number_input(
    "Uniform Depth (ft)",
    min_value=0.1,
    value=0.5,
    step=0.1,
    disabled=not use_uniform_depth,
    help="Flood depth applied everywhere when using the uniform option",
)
period_years = st.sidebar.number_input(
    "📆 Analysis Period (Years)",
    min_value=1,
    value=50,
    help="Used for context in planning studies; not required for EAD computation",
)
samples = st.sidebar.number_input(
    "🎲 Monte Carlo Iterations",
    min_value=10,
    value=100,
    help="Number of random simulations per crop type to estimate uncertainty",
)
depth_sd = st.sidebar.number_input(
    "± Depth Uncertainty (ft)",
    value=0.1,
    help="Assumed standard deviation of flood depth error (used only in Monte Carlo)",
)
value_sd = st.sidebar.number_input(
    "± Crop Value Uncertainty (%)",
    value=10,
    help="Assumed variability in per-acre crop values (used only in Monte Carlo)",
)
month_uncertainty = st.sidebar.checkbox(
    "📅 Randomize Flood Month",
    value=False,
    help="If checked, Monte Carlo simulations will sample flood months uniformly across the year",
)

crop_inputs, label_to_filename, label_to_metadata = {}, {}, {}

# Process cropland raster
if crop_file:
    crop_path = save_upload(crop_file, ".tif")
    st.session_state.crop_path = crop_path

    with rasterio.open(crop_path) as src:
        arr = src.read(1)
    counts = Counter(arr.flatten())
    codes = [c for c, _ in counts.most_common() if c != 0]

    st.markdown("### 🌱 Crop Values and Growing Seasons")
    st.caption(
        "Default values and calendars are provided. Expand below to adjust assumptions as needed."
    )
    with st.expander("Review or customize crop assumptions", expanded=False):
        for code in codes:
            # Use the crop code itself as a fallback name so undefined codes are
            # clearly identifiable in the UI instead of appearing blank.
            default_name, default_val = CROP_DEFINITIONS.get(code, (str(code), 0))
            default_season = CROP_GROWING_SEASONS.get(
                code, DEFAULT_GROWING_SEASON
            )
            name = st.text_input(
                f"Crop {code} – Name",
                value=default_name,
                key=f"name_{code}",
                help="Descriptive name for this crop code",
            )
            # ``default_val`` is cast to ``float`` above, so ``step`` must also be a
            # float. Use a small increment and explicit format so precise dollar
            # amounts (e.g., 1193.19) are preserved instead of being rounded.
            val = st.number_input(
                f"{name} – $/Acre",
                value=float(default_val or 0),
                step=0.01,
                format="%.2f",
                key=f"val_{code}",
                help="Enter average crop value per acre for this crop",
            )
            season = st.multiselect(
                f"{name} – Growing Months",
                list(range(1, 13)),
                default=default_season,
                key=f"season_{code}",
                help="Months when this crop is vulnerable to flooding",
            )
            crop_inputs[code] = {"Name": name, "Value": val, "GrowingSeason": season}
    st.session_state.crop_inputs = crop_inputs

# Process flood rasters or uniform depth
if depth_files or use_uniform_depth:
    st.markdown("### ⚙️ Flood Raster Settings")
    depth_inputs = []

    if depth_files:
        for i, f in enumerate(depth_files):
            path = save_upload(f, ".tif")
            label = sanitize_label(os.path.splitext(os.path.basename(f.name))[0])
            depth_inputs.append((label, path))
            rp = st.number_input(
                f"Return Period: {f.name}",
                min_value=1,
                value=100,
                key=f"rp_{i}",
                help="How often this flood event is expected to occur (e.g., 100 for 1-in-100 year flood)",
            )
            mo = st.number_input(
                f"Flood Month: {f.name}",
                min_value=1,
                max_value=12,
                value=6,
                key=f"mo_{i}",
                help="Month of flood to compare against crop growing season",
            )
            label_to_filename[label] = f.name
            label_to_metadata[label] = {"return_period": rp, "flood_month": mo}

    if use_uniform_depth and crop_file:
        rp = st.number_input(
            "Return Period: Uniform Depth",
            min_value=1,
            value=100,
            key="rp_uniform",
            help="How often this flood event is expected to occur",
        )
        mo = st.number_input(
            "Flood Month: Uniform Depth",
            min_value=1,
            max_value=12,
            value=6,
            key="mo_uniform",
            help="Month of flood to compare against crop growing season",
        )

        depth_arr = constant_depth_array(st.session_state.crop_path, uniform_depth_ft)
        label = sanitize_label(f"uniform_{uniform_depth_ft}ft")
        depth_inputs.append((label, depth_arr))
        label_to_filename[label] = label
        label_to_metadata[label] = {"return_period": rp, "flood_month": mo}


    st.session_state.depth_inputs = depth_inputs
    st.session_state.label_map = label_to_filename
    st.session_state.label_metadata = label_to_metadata

# Direct Damages Mode
if mode == "Direct Damages":
    if st.button("🚀 Run Flood Damage Estimator"):
        if not (crop_file and (depth_files or use_uniform_depth)):
            st.error(
                "❌ Please upload a cropland raster and at least one flood source."
            )
        else:
            cleanup_temp_dir()
            st.session_state.temp_dir = tempfile.TemporaryDirectory()
            with st.spinner("🔄 Processing flood damages..."):
                try:
                    result_path, summaries, diagnostics, _ = process_flood_damage(
                        st.session_state.crop_path,
                        st.session_state.depth_inputs,
                        st.session_state.temp_dir.name,
                        period_years,
                        crop_inputs,
                        st.session_state.label_metadata,
                    )
                    st.session_state.result_path = result_path
                    st.session_state.summaries = summaries
                    st.session_state.diagnostics = diagnostics
                    st.success("✅ Direct damage analysis complete.")
                except Exception as e:
                    st.error(f"❌ Error: {e}")

    if st.session_state.result_path:
        st.markdown("---")
        st.header("📈 Results & Visualizations")

        image_paths = {}

        for label, df in st.session_state.summaries.items():
            st.subheader(f"📋 Summary for {label}")
            with st.expander("ℹ️ Column Definitions"):
                st.markdown(
                    """
                - **CropCode**: CropScape numerical code.
                - **CropName**: Descriptive crop type.
                - **FloodedPixels**: Number of inundated pixels.
                - **FloodedAcres**: Area affected (pixels × pixel area; 1 pixel ≈ 0.222 acres).
                - **ValuePerAcre**: Input value per acre for the crop.
                - **DollarsLost**: Total crop damage.
                - **EAD**: Expected Annual Damage = DollarsLost ÷ ReturnPeriod.
                """
                )
            st.dataframe(df)
            chart_data = (
                df.sort_values("DollarsLost", ascending=False)
                .head(10)
                .set_index("CropName")["DollarsLost"]
            )
            fig_bar, ax_bar = plt.subplots()
            ax_bar.bar(chart_data.index.astype(str), chart_data.values)
            ax_bar.set_xlabel("Crop Type")
            ax_bar.set_ylabel("Dollars Lost")
            ax_bar.tick_params(axis="x", labelrotation=90)
            st.pyplot(fig_bar)
            bar_path = os.path.join(
                st.session_state.temp_dir.name, f"bar_{label}.png"
            )
            # Ensure the entire bar chart (including labels) is saved without cropping
            fig_bar.savefig(bar_path, bbox_inches="tight")
            plt.close(fig_bar)
            image_paths.setdefault(label, {})["bar"] = bar_path

        if st.session_state.diagnostics:
            st.subheader("🛠️ Diagnostics")
            st.dataframe(pd.DataFrame(st.session_state.diagnostics))

        # Insert images into Excel
        if image_paths:
            wb = load_workbook(st.session_state.result_path)
            for label, imgs in image_paths.items():
                if label in wb.sheetnames:
                    ws = wb[label]
                    if "bar" in imgs:
                        ws.add_image(XLImage(imgs["bar"]), "H2")
            wb.save(st.session_state.result_path)

        with open(st.session_state.result_path, "rb") as file:
            st.download_button(
                label="📥 Download Results Excel File",
                data=file,
                file_name=os.path.basename(st.session_state.result_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# Monte Carlo Mode
elif mode == "Monte Carlo Simulation":
    if st.button("🧪 Run Monte Carlo Simulation"):
        if not (crop_file and (depth_files or use_uniform_depth)):
            st.error(
                "❌ Please upload a cropland raster and at least one flood source."
            )
        else:
            cleanup_temp_dir()
            st.session_state.temp_dir = tempfile.TemporaryDirectory()
            with st.spinner("🔬 Running Monte Carlo..."):
                try:
                    result_path, summaries, diagnostics, _ = process_flood_damage(
                        st.session_state.crop_path,
                        st.session_state.depth_inputs,
                        st.session_state.temp_dir.name,
                        period_years,
                        crop_inputs,
                        st.session_state.label_metadata,
                    )
                    mc_results = run_monte_carlo(
                        summaries,
                        st.session_state.label_metadata,
                        samples,
                        value_sd,
                        depth_sd,
                        month_uncertainty,
                    )
                    st.session_state.result_path = result_path

                    st.markdown("---")
                    st.header("📉 Monte Carlo EAD Results")

                    for label, df in mc_results.items():
                        st.subheader(f"🧪 MC Summary for {label}")
                        with st.expander("ℹ️ Column Definitions"):
                            st.markdown(
                                """
                            - **CropCode**: CropScape numerical code.
                            - **CropName**: Descriptive crop type.
                            - **EAD_MC_Mean**: Mean simulated EAD from Monte Carlo.
                            - **EAD_MC_5th / 95th**: Uncertainty bounds (percentiles).
                            - **Original_EAD**: Deterministic EAD value for comparison.
                            """
                            )
                        st.dataframe(df)
                        chart_data = (
                            df.sort_values("EAD_MC_Mean", ascending=False)
                            .head(10)
                            .set_index("CropName")["EAD_MC_Mean"]
                        )
                        fig_mc, ax_mc = plt.subplots()
                        ax_mc.bar(chart_data.index.astype(str), chart_data.values)
                        ax_mc.set_xlabel("Crop Type")
                        ax_mc.set_ylabel("EAD_MC_Mean")
                        ax_mc.tick_params(axis="x", labelrotation=90)
                        st.pyplot(fig_mc)
                        plt.close(fig_mc)

                    with pd.ExcelWriter(
                        result_path, mode="a", engine="openpyxl"
                    ) as writer:
                        for label, df in mc_results.items():
                            df.to_excel(writer, sheet_name=f"MC_{label}", index=False)

                    with open(result_path, "rb") as file:
                        st.download_button(
                            label="📥 Download Monte Carlo Excel File",
                            data=file,
                            file_name=os.path.basename(result_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

                    st.success("✅ Monte Carlo analysis complete.")

                except Exception as e:
                    st.error(f"⚠️ Monte Carlo error: {e}")

